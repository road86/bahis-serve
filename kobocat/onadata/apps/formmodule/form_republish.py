import string
import openpyxl
from openpyxl.cell import get_column_letter
import psycopg2
#from django.contrib.auth.models import User
#from django.utils.crypto import get_random_string
from pyxform.builder import create_survey_from_xls
#from pyxform.xls2json import SurveyReader
from pyxform.xls2json_backends import xls_to_dict, csv_to_dict
import pyxform.constants as constants
import pyxform.aliases as aliases
import os
import re
from django.db import connection

import pandas as pd
import json

#connection = psycopg2.connect("dbname='coredbdada' user='kobo' host='192.168.19.89' password='DB@mPower@786'")
def _section_name(path_or_file_name):
    directory, filename = os.path.split(path_or_file_name)
    section_name, extension = os.path.splitext(filename)
    return section_name


def get_filename(path):
    """
    Get the extensionless filename from a path
    """
    return os.path.splitext((os.path.basename(path)))[0]


def parse_file_to_workbook_dict(path, file_object=None):
    """
    Given a xls or csv workbook file use xls2json_backends to create
    a python workbook_dict.
    workbook_dicts are organized as follows:
    {sheetname : [{column_header : column_value_in_array_indexed_row}]}
    """
    (filepath, filename) = os.path.split(path)
    if not filename:
        raise PyXFormError("No filename.")
    (shortname, extension) = os.path.splitext(filename)
    if not extension:
        raise PyXFormError("No extension.")

    if extension == ".xls" or extension == ".xlsx":
        return xls_to_dict(file_object if file_object is not None else path)
    elif extension == ".csv":
        return csv_to_dict(file_object if file_object is not None else path)
    else:
        raise PyXFormError("File was not recognized")


def parse_file_to_json(path, default_name=None, default_language=u"default",
                       warnings=None, file_object=None):
    """
    A wrapper for workbook_to_json
    """
    workbook_dict = parse_file_to_workbook_dict(path, file_object)
    #print(workbook_dict['survey'])
    if default_name is None:
        default_name = unicode(get_filename(path))
    #return workbook_to_json(workbook_dict, default_name, default_language, warnings)
    #return workbook_dict[constants.SURVEY]
    return workbook_dict


def parse_file_to_workbook_dict(path, file_object=None):
    """
    Given a xls or csv workbook file use xls2json_backends to create
    a python workbook_dict.
    workbook_dicts are organized as follows:
    {sheetname : [{column_header : column_value_in_array_indexed_row}]}
    """
    (filepath, filename) = os.path.split(path)
    if not filename:
        raise PyXFormError("No filename.")
    (shortname, extension) = os.path.splitext(filename)
    if not extension:
        raise PyXFormError("No extension.")

    if extension == ".xls" or extension == ".xlsx":
        return xls_to_dict(file_object if file_object is not None else path)
    elif extension == ".csv":
        return csv_to_dict(file_object if file_object is not None else path)
    else:
        raise PyXFormError("File was not recognized")

class SpreadsheetReader(object):
    def __init__(self, path_or_file):
        path = path_or_file
        if type(path_or_file) is file:
            path = path.name
        self._dict = parse_file_to_workbook_dict(path)
        self._path = path
        self._id = unicode(get_filename(path))
        self._name = self._print_name = self._title = self._id

    def to_json_dict(self):
        return self._dict

    #TODO: Make sure the unicode chars don't show up
    def print_json_to_file(self, filename=""):
        if not filename:
            filename = self._path[:-4] + ".json"
        print_pyobj_to_json(self.to_json_dict(), filename)


class SurveyReader(SpreadsheetReader):
    """
    SurveyReader is a wrapper for the parse_file_to_json function.
    It allows us to use the old interface where a SpreadsheetReader
    based object is created then a to_json_dict function is called on it.
    """
    def __init__(self, path_or_file):
        if isinstance(path_or_file, basestring):
            self._file_object = None
            path = path_or_file
        else:
            self._file_object = path_or_file
            path = path_or_file.name

        self._warnings = []
        self._dict = parse_file_to_json(
            path, warnings=self._warnings, file_object=self._file_object)
        self._path = path

    def print_warning_log(self, warn_out_file):
        #Open file to print warning log to.
        warn_out = open(warn_out_file, 'w')
        warn_out.write('\n'.join(self._warnings))

    def get_survey_dictionary(self):
        #Open file to print warning log to.        
        return self._dict


def get_form_xls(path,new_path):
    name = _section_name(path)

    excel_reader = SurveyReader(path)
    form_dic= excel_reader.get_survey_dictionary()
    survey_dic= form_dic[constants.SURVEY]

    if constants.CHOICES in form_dic.keys():
        choice_dic= form_dic[constants.CHOICES]
    else:
        choice_dic={}
    
    if (constants.CHOICES+'_header') in form_dic.keys():
        choice_dic_header= form_dic[constants.CHOICES+'_header']
    else:
        choice_dic_header={}    
    
    #choice_dic= form_dic[constants.CHOICES]
    #choice_dic_header= form_dic[constants.CHOICES+'_header']
    
    setting_dic = form_dic[constants.SETTINGS]
        
    form_id = setting_dic[0]["form_id"]



    select_regexp = re.compile(
            r"^(?P<select_command>(" + '|'.join(aliases.multiple_choice.keys())
            + r")) (?P<list_name>\S+)"
            + "( (?P<specify_other>(or specify other|or_other|or other)))?$")
    print(select_regexp)
    master_list_names=[]
    master_dict_names={}
    search_list_names=[]
    for field in survey_dic:
        
        if 'type' in field:
            #print(field['type'])
            question_type = field['type']
            select_parse = select_regexp.search(question_type)
            if select_parse:
                
                parse_dict = select_parse.groupdict()
                list_name=''
                if 'list_name' in parse_dict:
                    list_name = parse_dict['list_name']
                
                field =  {k.lower(): v for k, v in field.items()}

                
                
                search=""
                if 'appearance' in field.keys():
                    appearance = field['appearance']
                    #x = re.findall("search[ ]*[(][ ]*'[a-zA-Z]+[a-zA-Z_0-9]*'[ ]*[)]", appearance)
                    x = re.findall("search[ ]*[(][' ,{}$a-zA-Z_0-9]*[)]", appearance)
                    if list_name == 'schedule':
                        print(appearance)
                        print(len(x))
                    #search=""
                    if len(x)>0:
                        search = x[0]
                    #print(search)
                    if list_name == 'schedule':
                        print(search)
                        
                choice_filter=""
                if 'choice_filter' in field.keys():
                    choice_filter = field['choice_filter']
                    choice_filter = choice_filter.split("=")
                    choice_filter=choice_filter[0].strip()                        

                if (list_name != u"") and search =="" and list_name not in master_list_names:
                    master_list_names.append(list_name)
                    if list_name in master_dict_names:
                        if choice_filter !="":
                            master_dict_names[list_name].add(choice_filter)                        
                    else:
                        l=set()
                        if choice_filter !="":
                            l.add(choice_filter)
                        master_dict_names[list_name]=l
                                            
                elif (list_name != u"") and search !="" and list_name not in search_list_names:
                    if list_name == 'schedule':
                        print("tt"+search)
                    search_list_names.append(list_name)

                #print(parse_dict,list_name)
    #print(master_list_names)
    #print(search_list_names)
    #print (choice_dic)
    df_choice_sheet = pd.DataFrame()
    if len(choice_dic_header)>0:
        cdh = choice_dic_header[0]
        cdh =(cdh.keys())
        df_choice_sheet = pd.DataFrame(columns=cdh)
        #print(df_choice_sheet)

    #filter(lambda person: person['list_name'] == 'report_type', choice_dic)
    search_choice=[]
    for l in search_list_names:
        #d = filter(lambda sl: sl['list_name]' == l, choice_dic)
        #d={}
        #d['list_name']=l
        rows = filter(lambda sl: l == sl['list_name'] if 'list_name' in sl else '' , choice_dic)
        df = pd.DataFrame(rows)
        df_choice_sheet = df_choice_sheet.append(df,ignore_index=True)
        #print(df)
        
        #search_choice.append(rows)


    master_qry = """select mc.id catergory_id,category_name,coalesce(parent_id,0) as parent_id, 
                mci.id item_id,value, name_eng, name_bangla,coalesce(parent_item_id,0) as parent_item_id
                from core.master_category mc inner join core.master_category_item mci
                on mc.id=mci.category_id
                where mc.active = true ;"""
    master_df = pd.read_sql(master_qry,connection)

    master_df = pd.read_sql(master_qry,connection)

    master_parent_df=pd.DataFrame()
    master_parent_df[["parent_item_id","parent_value"]]=master_df[["item_id","value"]]

    master_df = master_df.merge(master_parent_df,how="left",on=["parent_item_id"])
    
    
    master_df["parent_item_id"]=master_df["parent_item_id"].fillna(value='')
    master_df["parent_id"]=master_df["parent_id"].fillna(value='')
    #print(choice_dic_header)


    for l in master_list_names:
        #print(l)
        df=master_df[master_df["category_name"]==l]
        #print(df)
        if len(df)<=0:
            rows = filter(lambda sl: l == sl['list_name'] if 'list_name' in sl else '' , choice_dic)
            df_choice = pd.DataFrame(rows)
        else:
            filter_list = master_dict_names[l]
            
            df_choice = pd.DataFrame()
            df_choice[["list_name","name","label::Bangla","label::English"]]=df[["category_name", "value","name_bangla","name_eng"]]
            for fl in list(filter_list):
                df_choice[[fl]] = df[["parent_value"]]        
        df_choice_sheet = df_choice_sheet.append(df_choice,ignore_index=True)
        #print(df_choice)
        #check whether exists in master data
    df_choice_sheet = df_choice_sheet.fillna('')
    #print(df_choice_sheet)    

    wb_obj = openpyxl.load_workbook(path) 

    print(wb_obj.get_sheet_names())
    
    try:
        ws_choice = wb_obj.get_sheet_by_name("choices")
        if len(df_choice_sheet)>0:
            wb_obj.remove_sheet(ws_choice)
    except:
        pass

    #wb_obj.create_sheet("choices")
    if len(df_choice_sheet)>0:
        ws_choice = wb_obj.create_sheet(index= 0 ,title="choices") # insert at first position
        #ws_choice = wb_obj.get_sheet_by_name("choices")

        cols = list(df_choice_sheet)
        i = 1

        for l in cols:
            col = get_column_letter(i)
            i = i + 1
            ws_choice.cell('%s%s'%(col, 1)).value = '%s' % (l)

        row=2
        for ind,rw in df_choice_sheet.iterrows():
            i=1
            for l in cols:
                col = get_column_letter(i)
                i = i + 1
                val = rw[l]
                ws_choice.cell('%s%s'%(col, row)).value = '%s' % (val)
            row = row + 1

    wb_obj.save(new_path)







# def get_first_non_empty_row(ws):
#     return 1
#
# def get_master_data(master_data_name):
#     file_name=master_data_name+".csv"
#     qry="""select mc.id,mc.category_name,mci.value,mci.name_eng,mci.name_bangla,mci.parent_item_id  from master_category mc
#            left outer join core.master_category_item mci on mc.id = mci.category_id where mc.category_name='%s'""" %master_data_name
#     df=pd.read_csv(file_name)
#     return df
#
# def upload_bulk_member():
#     print("in celery process")
#
#     #print (settings.MEDIA_ROOT)
#     user_path_filename="farm_assessment.xlsx"
#
#     wb_obj = openpyxl.load_workbook(user_path_filename)
#     print(wb_obj.sheetnames)
#
#     ws_survey = wb_obj.get_sheet_by_name("survey")
#     max_row= ws_survey.max_row
#     max_column = ws_survey.max_column
#
#     first_not_empty_row = get_first_non_empty_row(ws_survey)
#
#     col_names = []
#     i=0
#     type_col = -1
#     choice_col = -1
#
#     for column in ws_survey.iter_cols(first_not_empty_row, ws_survey.max_column):
#         col_names.append(column[0].value)
#         if str(column[0].value)=="type":
#             type_col = i
#         elif str(column[0].value)=="choice_filter":
#             choice_col = i
#         i = i + 1
#
#     print(type_col)
#
#     select_list=list()
#     for row in ws_survey.iter_rows(max_row=ws_survey.max_row):
#         type_val = str(row[type_col].value)
#         choice_filter = str(row[choice_col].value) if choice_col != -1 else ""
#         if (type_val[:10]).upper()=="SELECT_ONE":
#             type_val = type_val.split(' ')
#             if len(type_val)>1:
#                 type_val = type_val[1]
#                 choice_filter = choice_filter.split("=")
#                 choice_filter = choice_filter[0]
#                 d={}
#                 d["list_name"]=type_val
#                 d["choice_filter"]=choice_filter
#                 select_list.append(d)
#
#     print(select_list)
#
#     try:
#         ws_choice = wb_obj.get_sheet_by_name("choices")
#     except:
#         pass
#     else:
#         wb_obj.remove(ws_choice)
#     wb_obj.create_sheet("choices")
#     ws_choice = wb_obj.get_sheet_by_name("choices")
#
#     languages=['Bangla','English']
#
#     i = 65
#     ws_choice[str(chr(i)) + "1"] = "list_name"
#     i = i + 1
#     ws_choice[str(chr(i)) + "1"] = "name"
#
#     for l in languages:
#         i = i + 1
#         ws_choice[str(chr(i)) + "1"] = "label::" + l
#
#     i = i + 1
#     ws_choice[str(chr(i)) + "1"] = "myfilter"
#     i = i + 1
#     ws_choice[str(chr(i)) + "1"] = "image"
#     i = i + 1
#     ws_choice[str(chr(i)) + "1"] = "audio"
#     i = i + 1
#     ws_choice[str(chr(i)) + "1"] = "video"
#
#     r_number = 2
#
#     for sl in select_list:
#         if sl["list_name"] == "report_type":
#             df=get_master_data(sl["list_name"])
#             for id,row in df.iterrows():
#                 i = 65
#                 ws_choice[chr(i) + str(r_number)] = row["list_name"]
#                 i += 1
#                 ws_choice[chr(i) + str(r_number)] = row["name"]
#                 for l in languages:
#                     i += 1
#                     ws_choice[chr(i) + str(r_number)] = row["label::" + l]
#                 i += 1
#                 ws_choice[chr(i) + str(r_number)] = row["myfilter"]
#                 r_number +=1
#
#
#
#
#
#
#     wb_obj.save("a.xlsx")
